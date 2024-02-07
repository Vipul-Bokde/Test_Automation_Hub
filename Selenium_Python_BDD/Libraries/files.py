import base64
import os
import glob
import csv
import re
import time
import logging
import json
import zipfile

from datetime import date
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from libraries import cnp_logs

LOGGER = cnp_logs.get_logger("files")


# --------------------
# Read files Functions
# --------------------


def get_download_path():
    """Returns the default uploadfiles path"""
    return '/'.join(os.getcwd().split('/')[:3]) + '/Downloads'


def get_last_file_downloaded():
    """Returns path of the last file downloaded"""
    time.sleep(3)
    list_of_files = glob.glob(get_download_path() + "/*")  # means all if need specific format then .csv
    latest_file = max(list_of_files, key=os.path.getctime)
    return latest_file


def get_string_desired(string, text_desired):
    """Returns a string in another string"""
    temp1 = string.find(text_desired)
    if temp1 != -1:
        length = len(string)
        text = string[temp1:length]
        return text
    else:
        return "Text not found"


def get_root_directory():
    return os.path.dirname(os.path.realpath(__import__("__main__").__file__))


def verify_last_file_downloaded(file_name):
    """Compare the last file name with a given name"""
    today = date.today()
    file_name = file_name + str(today)
    file_path = get_last_file_downloaded()
    name = get_string_desired(file_path, file_name)
    length = len(file_name)
    name_new = name[:length]
    if file_name in name_new:
        return True
    else:
        return False


# --------------------
# Read files from Remote
# --------------------


def get_downloaded_files(self):
    files = self.driver.get("chrome://downloads")
    return self.driver.execute_script(
        "return  document.querySelector('downloads-manager')  "
        " .shadowRoot.querySelector('#downloadsList')         "
        " .items.filter(e => e.state === 'COMPLETE')          "
        " .map(e => e.filePath || e.file_path || e.fileUrl || e.file_url); ")


def get_file_content(self, path):
    try:
        elem = self.driver.execute_script(
            "var input = window.document.createElement('INPUT'); "
            "input.setAttribute('type', 'file'); "
            "input.hidden = true; "
            "input.onchange = function (e) { e.stopPropagation() }; "
            "return window.document.documentElement.appendChild(input); ")
        elem._execute('sendKeysToElement', {'value': [path], 'text': path})
        result = self.driver.execute_async_script(
            "var input = arguments[0], callback = arguments[1]; "
            "var reader = new FileReader(); "
            "reader.onload = function (ev) { callback(reader.result) }; "
            "reader.onerror = function (ex) { callback(ex.message) }; "
            "reader.readAsDataURL(input.files[0]); "
            "input.remove(); "
            , elem)
        if not result.startswith('data:'):
            raise Exception("Failed to get file content: %s" % result)
        return base64.b64decode(result[result.find('base64,') + 7:])
    finally:
        logging.info("get_file_content executed successfully")


def copy_remote_file(self):
    ENV = os.environ.get("SPARC_SERVICE_NAME", None)
    files = get_downloaded_files(self)
    if files:
        content = get_file_content(self, files[0])
        file_basename = os.path.basename(files[0])
        download_file_loc = os.path.join(ENV + "/automation_test/downloaded_files", file_basename)
        with open(download_file_loc, 'wb') as f:
            f.write(content)
        return download_file_loc


# -------------------
# CSV file functions
# -------------------

def read_csv_file(file_path):
    """
        Author : Luz A. Vargas
        Description : This method opens and reads a .csv file
        Arguments : file_path: String e.g. 'UBR/automation_test/downloaded_files/ESI MEDICAL 2022Q3-calc.csv'
        Returns : a reader object
    """
    file = open("." + file_path)
    reader = csv.reader(file, delimiter=',')
    return reader


def read_headers_from_csv_file(file_path):
    """
        Author : Luz A. Vargas
        Description : This method opens and reads a .csv file
        Arguments : file_path: String e.g. 'UBR/automation_test/downloaded_files/ESI MEDICAL 2022Q3-calc.csv'
        Returns : a dictreader object
    """
    file = open(file_path)
    reader = csv.DictReader(file)
    return reader


def get_values_from_column(file_data, column_name):
    """
        Author : Luz A. Vargas
        Description : This method gets the column values from a .csv file
        Arguments : file_data: dictreader object, column_name: String e.g. 'Market Share Amount'
        Returns : a list of the column values
    """
    col_values = []
    for col in file_data:
        cell_value = float(col[column_name])
        if cell_value != float(0):
            col_values.append(cell_value)
    LOGGER.info("col_values_calc: " + str(col_values))
    return col_values


def find_something_in_csv_file(path, items_to_search):
    result = False
    with open(path) as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        for row in csv_reader:
            for item in items_to_search:
                if str(row).find(item) != -1:
                    result = True
    return result


def compare_headers_in_csv_file(path_csv, first_column):
    """Compare the csv file headers with a list and return the result"""
    result = False
    with open(path_csv) as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        first_row = 0
        for rows in csv_reader:
            if first_row == 0:
                for x in range(len(rows)):
                    if rows[x] != first_column[x]:
                        print("Column different in the csv file: ", rows[x])
                        result = False
                        break
                    else:
                        result = True
            break
    return result


def get_list_through_a_string(string, symbol):
    """Return a list with info separated by commas"""
    list_with_info = string.split(symbol)
    return list_with_info


# ------------------------
# Create and Delete files
# ------------------------

def create_file(file_path, file_body):
    f = open("." + file_path, "w")
    f.write(file_body)
    f.close()


def delete_file(file_path):
    if os.path.isfile("." + file_path):
        os.remove("." + file_path)
    else:
        LOGGER.error("File not found")


# ------------------------
# Read & Write txt files
# ------------------------

def write_into_txt_as_KeyPair(file_Path, str_Key, str_Value):
    details = {str_Key: str_Value}
    with open(os.getcwd() + '\\Medicaid\\TestData\\' + file_Path, 'w') as f:
        f.write(json.dumps(details))


def readFile_as_key_value_pair(file_path, str_key):
    f = open(os.getcwd() + '\\Medicaid\\TestData\\' + file_path, 'r')
    line = f.read()
    js = json.loads(line)
    f.close()
    return js[str_key]


def clear_text_file(file_path):
    with open(os.getcwd() + '\\Medicaid\\TestData\\' + file_path, 'w') as f:
        f.close()


# -------------------
# Excel file functions
# -------------------

def open_an_excel_file(file_name):
    """
        Author : Luz A. Vargas
        Description : This method opens an Excel and gets the names of the worksheets
        Arguments : file_name: String e.g. 'UBR/automation_test/downloaded_files/rebate_recon_ferring_17700.zip'
        Returns : a dictionary of the workbook and a list of the worksheets contained in the Excel file
    """
    file = load_workbook(filename=file_name)
    sheets = file.sheetnames
    return {"file": file, "sheets": sheets}


def get_columns_from_excel(file_name):
    """
        Author : Luz A. Vargas
        Description : This method extracts from the Excel file the names of the columns included in the worksheets
        Arguments : file_name: String e.g. 'UBR/automation_test/downloaded_files/rebate_recon_ferring_17700.zip'
        Returns : a dictionary of the columns that each worksheet contains
    """
    wb = open_an_excel_file(file_name)
    wb_data = {}
    for sheet_name in wb["sheets"]:
        ws = wb["file"][sheet_name]
        header_row = ws[1]
        columns = [cell.value for cell in header_row]
        wb_data[sheet_name] = columns
    return wb_data


def find_a_specific_cell(ws, column_name):
    """
        Author : Luz A. Vargas
        Description : This method loops through the rows and columns of a worksheet in an Excel file
        Arguments : ws: Worksheet object, column_name: String e.g. 'Market Share'
        Returns : a target cell
    """
    for row in range(1, ws.max_row + 1):
        for column in range(1, ws.max_column + 1):
            col = get_column_letter(column)
            cell_name = f'{col}{row}'
            cell_value = ws[cell_name].value
            if cell_value is not None and isinstance(cell_value, str):
                if cell_value.strip() == column_name.strip():
                    return cell_name


def sum_values_list(column_values):
    """
        Author : Luz A. Vargas
        Description : This method sums the cell values provided in a list on a worksheet in an Excel file
        Arguments : column_values: List e.g. [1.9, 5.6]
        Returns : the sum of the cell values
    """
    new_column_values = [float(value) for value in column_values if value != 'None']
    new_cell_value = round(sum(new_column_values), 2)
    if new_cell_value is None or new_cell_value == 0.0:
        new_cell_value = 0.00
    return new_cell_value


def get_cell_value_from_formula(cell_value, wb, column_values, sum_values):
    """
        Author : Luz A. Vargas
        Description : This method gets the cell value from a formula in a worksheet in an Excel file
        Arguments : cell_value: String e.g. '=SUM(A4:A23)', wb: Workbook object, column_values: List e.g. [1.9, 5.6],
        sum_values: List e.g. [1.9, 5.6]
        Returns : a float or a string depending on the provided cell value
    """
    for sheet in wb["sheets"]:
        if sheet in cell_value:
            ws = wb["file"][sheet]
            cell_name = cell_value.split("!")[1]
            new_cell_value = ws[cell_name].value
            if new_cell_value is None:
                new_cell_value = "0"
            return new_cell_value
        if "SUM" in cell_value:
            if sum_values is not None:
                column_values = sum_values
            new_cell_value = sum_values_list(column_values)
            return new_cell_value


def check_total_cell_by_worksheet(row, ws):
    """
        Author : Luz A. Vargas
        Description : This method checks if the cell value is the Total in a worksheet in an Excel file
        Arguments : row: int e.g. 5, ws: Worksheet object
        Returns : a Boolean depending on whether if the cell is the Total
    """
    check_cell = f'A{row}'
    check_cell_value = str(ws[check_cell].value)
    if "Total" in check_cell_value or "Totals" in check_cell_value or "Amount Paid" in check_cell_value:
        return True
    return False


def capture_total_and_column_values(column_name, cell_value, column_values=None):
    """
        Author : Luz A. Vargas
        Description : This method capture the total and column values in a worksheet in an Excel file
        Arguments : column_name: String e.g. 'Market Share', cell_value: String e.g. '5.94',
        column_values: List e.g. [1.9, 5.6]
        Returns : a dictionary of the target column values and the Total
    """
    col_values = {}
    col_name = column_name + " Total"
    if isinstance(cell_value, str):
        cell_value = float(cell_value)
    col_values[col_name] = cell_value
    col_values[column_name] = column_values
    return col_values


def get_value_from_a_specific_cell(ws, column_name, target_column, row_totals):
    """
        Author : Luz A. Vargas
        Description : This method gets the value for a specific cell of a worksheet in an Excel file
        Arguments : ws: Worksheet object, column_name: String e.g. 'Market Share', target_column: String e.g. 'A1',
        row_totals: int e.g. 125
        Returns : the cell value in a list type
    """
    t_column = re.split(r'(\d+)', target_column)
    cell_name = f'{t_column[0]}{row_totals}'
    cell_value = str(round(ws[cell_name].value, 2))
    col_values = capture_total_and_column_values(column_name, cell_value)
    return col_values


def get_values_by_cell(ws, target_column, column_name, wb, sum_values):
    """
        Author : Luz A. Vargas
        Description : This method gets the cell values through a column of a worksheet in an Excel file
        Arguments : ws: Worksheet object, target_column: String e.g. 'A1', column_name: String e.g. 'Market Share',
        wb: Workbook object, sum_values: float e.g. 37226.32
        Returns : a dictionary of the target column values
    """
    column_values = []
    t_column = re.split(r'(\d+)', target_column)
    last_row = ws.max_row
    for row in range(1, last_row + 1):
        for column in t_column[0]:
            if row >= int(t_column[1]):
                cell_name = f'{column}{row}'
                cell_value = ws[cell_name].value
                if isinstance(cell_value, int):
                    cell_value = str(cell_value)
                if isinstance(cell_value, float):
                    cell_value = str(round(cell_value, 2))
                if cell_value is not None and column_name not in cell_value:
                    if "=" in str(cell_value):
                        while "=" in str(cell_value):
                            cell_value = get_cell_value_from_formula(cell_value, wb, column_values, sum_values)
                            if "=" not in str(cell_value) and "Cover Sheet" in str(ws) or "Payment Summary" in str(ws):
                                if check_total_cell_by_worksheet(row, ws):
                                    col_values = capture_total_and_column_values(column_name, cell_value, column_values)
                                    return col_values
                    elif row == last_row:
                        col_values = capture_total_and_column_values(column_name, cell_value, column_values)
                        col_values["row_totals"] = row
                        return col_values
                    column_values.append(cell_value)
                elif cell_value is None and "Cover Sheet" not in str(ws) or "Payment Summary" not in str(ws):
                    cell_value = 0
                    column_values.append(cell_value)


def get_values_in_worksheet_by_column(file_name, target_ws, column_name, sum_values=None, totals_row=None):
    """
        Author : Luz A. Vargas
        Description : This method gets the values of a given column, in a worksheets of the Excel file
        Arguments : file_name: String e.g. 'UBR/automation_test/downloaded_files/rebate_recon_ferring_17700.zip',
        target_ws: String e.g. 'Cover Sheet', column_name: String e.g. 'Market Share', sum_values: float e.g. 37226.32,
        row_totals: int e.g. 125
        Returns : the values for a target column
    """
    sheet_data = {}
    wb = open_an_excel_file(file_name)
    for sheet_name in wb["sheets"]:
        ws = wb["file"][sheet_name]
        if target_ws == sheet_name:
            target_column = find_a_specific_cell(ws, column_name)
            if target_column is not None:
                if totals_row is not None:
                    column_values = get_value_from_a_specific_cell(ws, column_name, target_column, totals_row)
                else:
                    column_values = get_values_by_cell(ws, target_column, column_name, wb, sum_values)
                sheet_data[target_ws] = column_values
            else:
                return False
    return sheet_data


def check_next_column(ws, current_column, target_columns):
    """
        Author : Luz A. Vargas
        Description : This method checks the column next to the current column, in a worksheets of the Excel file
        Arguments : ws: Worksheet object, current_column: String e.g. 'A1', target_columns: String e.g. 'Total Payment'
        or List e.g. ['Total Payment', 'Market Share']
        Returns : a Boolean depending on whether the target column is next to the current column
    """
    xy = coordinate_from_string(current_column)
    col = column_index_from_string(xy[0])
    new_column = get_column_letter(col + 1)
    cell_name = f'{new_column}{current_column[1:]}'
    t_column = ws[cell_name].value
    if not isinstance(target_columns, list):
        target_columns = [target_columns]
    for column in target_columns:
        if t_column in column:
            return True
    return False


def check_column_in_sheet(file_name, target_ws, column_name, target_columns=None):
    """
        Author : Luz A. Vargas
        Description : This method check that a column is in a worksheets of the Excel file
        Arguments : file_name: String e.g. 'UBR/automation_test/downloaded_files/rebate_recon_ferring_17700.zip',
        target_ws: String e.g. 'Cover Sheet', column_name: String e.g. 'Market Share',
        target_columns: String e.g. 'Total Payment' or List e.g. ['Total Payment', 'Market Share']
        Returns : True or False when target_ws is "Cover Sheet" or, "Not found" in otherwise
    """
    wb = open_an_excel_file(file_name)
    for sheet_name in wb["sheets"]:
        ws = wb["file"][sheet_name]
        if target_ws in sheet_name:
            current_column = find_a_specific_cell(ws, column_name)
            if current_column:
                if target_ws in "Cover Sheet" and column_name == "Market Share":
                    response = check_next_column(ws, current_column, target_columns)
                    return response
                else:
                    return True
            return "Not found"


def compare_excel_files(excel_file_path_one, excel_file_path_two):
    """Compare the Excel file headers with a list and return the result"""
    file1 = load_workbook(filename=excel_file_path_one)
    file2 = load_workbook(filename=excel_file_path_two)
    sheet1 = file1.sheetnames
    sheet2 = file2.sheetnames
    if sheet1 == sheet2:
        print("Both files have same sheet names")
    else:
        print("Sheet names in both files are different")
    compare_result = False
    for sheet in sheet1:
        ws1 = file1[sheet]
        ws2 = file2[sheet]
        for row in range(1, ws1.max_row + 1):
            for col in range(1, ws1.max_column + 1):
                if ws1.cell(row=row, column=col).value != ws2.cell(row=row, column=col).value:
                    print(f"Data in cell {ws1.cell(row=row, column=col).coordinate} is different")
                    compare_result = False
                    break
                else:
                    compare_result = True


"""
Author : Rushikesh Thakare
Description : This method navigate user to previous page
Arguments : NA
Returns : NA
"""


def navigate_to_back_page(self):
    self.driver.back()
    self.driver.refresh()


# -------------------
# Extract zip file functions
# -------------------

def extract_files_from_zip(file_name):
    """
        Author : Luz A. Vargas
        Description : This method extracts the contents of a .zip file, based on the file name passed with its path
        Arguments : file_name: String e.g. 'UBR/automation_test/downloaded_files/rebate_recon_ferring_17700.zip'
        Returns : a list of the files extracted from the .zip file
    """
    # opening the zip file in READ mode
    unzipped_files = []
    with zipfile.ZipFile(file_name, 'r') as zip:
        for info in zip.infolist():
            file = info.filename
            # extracting all the files
            unzipped_files.append(file)
        zip.extractall(path=file_name.split(file_name.split('/')[3])[0])
        return unzipped_files


# ------------------------------
# read and write into json file
# ------------------------------
"""
Author : Sadiya Kotwal
Description : This method write into json file
Arguments : service_name(service_name='GP' or 'Medicaid')
            file_name(file_name='test_data_read_write_jsonfile.json' )
            key(key='LabelerCode')
            Value(Value= '69990')
Returns : NA
"""


def write_into_json_file(service_name, file_name, key, Value):
    # Writing to sample.json
    with open(os.getcwd() + '\\' + service_name + '\\TestData\\JsonFiles\\' + file_name, "r") as jsonFile:
        data = json.load(jsonFile)
    data[key] = Value
    with open(os.getcwd() + '\\' + service_name + '\\TestData\\JsonFiles\\' + file_name, "w") as jsonFile_write:
        data = json.dump(data, jsonFile_write)


"""Author : Sadiya Kotwal
Description : This method reads from json file
Arguments : service_name(service_name='GP' or 'Medicaid')
            file_name(file_name='test_data_read_write_jsonfile.json' )
            key(key='LabelerCode')
Returns : data[Key](data[Key]='Value' Returns the value of key provided from json file)
"""


def read_from_json_file(service_name, file_name, Key):
    with open(os.getcwd() + '\\' + service_name + '\\TestData\\JsonFiles\\' + file_name, "r") as jsonFile:
        data = json.load(jsonFile)
        return data[Key]
