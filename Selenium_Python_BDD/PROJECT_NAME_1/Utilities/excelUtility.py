from xlrd import open_workbook
import openpyxl
from openpyxl import load_workbook
from time import sleep
import logging
import pandas
import json
import GP.utilities.Repo as Repo
import allure


def dataReader(filename,SheetName,TestCaseID,SearchValue):
        searchedString = None
        book = open_workbook(filename)
        sheet = book.sheet_by_name(SheetName)
        keys = [sheet.cell(0,col_index).value for col_index in range(sheet.ncols)]
        dict_list = []
        logging.info("No of rows in Excel", sheet.nrows)
        for row_index in range(0,sheet.nrows):
            for col_index in range(0,sheet.ncols):
                if sheet.cell_value(row_index,col_index) == TestCaseID:
                    testCaseNo = sheet.cell_value(row_index,col_index)
                    strRowNum = row_index

                    for col_index in range(0,sheet.ncols):
                        str1 = sheet.cell_value(0,col_index)
                        if str1 == SearchValue:
                            searchedString = sheet.cell_value(strRowNum, col_index)
                    return searchedString
                    break


def dataWriter(filename,TestSheetName,TestCaseID,ColumnName,ValueToStore):
        searchedString = None
        book = open_workbook(filename)
        sheet = book.sheet_by_name(TestSheetName)
        keys = [sheet.cell(0,col_index).value for col_index in range(sheet.ncols)]
        dict_list = []
        for row_index in range(0, sheet.nrows):
            for col_index in range(0, sheet.ncols):
                if sheet.cell_value(row_index, col_index) == TestCaseID:
                    testCaseNo = sheet.cell_value(row_index, col_index)
                    logging.info("Executing TestCase:-",testCaseNo)
                    strRowNum = row_index
                    logging.info("strRowNum:-",strRowNum)
                    for col_index in range(0, sheet.ncols):
                        str1 = sheet.cell_value(0,col_index)
                        if str1 == ColumnName:
                            wb = openpyxl.load_workbook(filename)
                            logging.info(wb.sheetnames)
                            sheetn= wb[TestSheetName]

                            c1= sheetn.cell(row= int(strRowNum)+1, column = int(col_index)+1)
                            c1.value = ValueToStore
                            sleep(1)
                            wb.save(filename)
                            logging.info("Writing Completed")
                        # if str1 == SearchValue:
                        #     searchedString = sheet.cell_value(strRowNum, col_index)
                        #     logging.info(searchedString)
                    return searchedString

def rowCountFetch(filename,SheetName):
        book = open_workbook(filename)
        sheet = book.sheet_by_name(SheetName)
        row_count = sheet.nrows
        # Subtract 1 from row count as excluding header
        row_count = row_count-1
        return  row_count

# Initialize data sets and read data
def dataReader_Initialize(SheetName,ScenarioID, TestID):
    searchedString = None
    excel_data_df = pandas.read_excel(Repo.testDataSheet_GP, sheet_name=SheetName,dtype=str)
    json_str = excel_data_df.to_json()
    json_str = excel_data_df.to_json(orient='records')
    loaded_json = json.loads(json_str)
    for x in loaded_json:
        if x['SCENARIO_ID'] == ScenarioID:
            if x['TC_ID'] == TestID:
                return x if x['EXECUTE'].lower() == 'y' else 'Since execution flag is \'''n''\' this data set will not be executed'


"""Author : Sadiya Kotwal
       Description : This open the excel sheet and print the sheet names in aluure
       Arguments : filepath(filepath='Alfasigma USA, Inc. GP Calc Audit Log 2023-07-31')
       Returns : NA""" 
def open_excel_and_print_sheetnames(filepath):
     # Load the entire workbook.
    obj_workbook = load_workbook(filepath)
    for sheetname in obj_workbook.sheetnames:
     allure.attach("Sheet names are : "+sheetname,attachment_type=allure.attachment_type.TEXT)


"""Author : Sadiya Kotwal
       Description : This method access the excel data frm specific row with merged columns
       Arguments : filepath(filepath='Alfasigma USA, Inc. GP Calc Audit Log 2023-07-31')
                sheetname(Eg: sheetname='C7R Report')
                start_row_index(Eg:start_row_index='0 or 1 ')
                end_row_index(end_row_index='2 to any index')
                specific_row_index(specific_row_index='3' or any row index whose data is to be acccessed)
               Returns : obj_specific_row(obj_specific_row='Returns the value of specific row whose data to be accessed')""" 
def access_data_from_Specific_row_and_for_merged_columns(filepath,sheetname,start_row_index,end_row_index,specific_row_index):
    #load workbook
    wb = load_workbook(filepath,data_only=True)
    #load worsheet
    obj_worksheet = wb[sheetname]
    # Count of max row
    obj_max_row_of_sheet =   obj_worksheet.max_row
    allure.attach("Maximum Rows that have data in sheet is : "+str(obj_max_row_of_sheet),attachment_type=allure.attachment_type.TEXT)
    # cOunt of Max sheet
    obj_max_column_of_sheet = obj_worksheet.max_column 
    allure.attach("Maximum Columns that have data in sheet is : "+str(obj_max_column_of_sheet),attachment_type=allure.attachment_type.TEXT)
    all_rows = list(obj_worksheet.rows)
    # This loop gives the data from the exact row which we need to access data
    for row in all_rows[start_row_index:end_row_index]:     # 1 to 5
        obj_specific_row = row[specific_row_index].value
    return obj_specific_row


"""Author : Sadiya Kotwal
       Description : This method access the excel data from single specific row and multiple columns from that row
       Arguments : filepath(filepath='Alfasigma USA, Inc. GP Calc Audit Log 2023-07-31')
                sheetname(Eg: sheetname='C7R Report')
                specific_row_index(specific_row_index='3' or any row index whose data is to be acccessed)
               Returns : list_of_single_row_multiple_columns(list_of_single_row_multiple_columns='Returns the list of the specific row and all columns of that row')""" 
def access_data_from_single_specific_row_and_multiple_columns_for_that_specific_row(filepath,sheetname,specific_row_index):
    wb = load_workbook(filepath,data_only=True)
    #get specific worksheet name
    obj_worksheet = wb[sheetname]
    obj_max_row_of_sheet =   obj_worksheet.max_row
    allure.attach("Maximum Rows that have data in sheet is : "+str(obj_max_row_of_sheet),attachment_type=allure.attachment_type.TEXT)
    obj_max_column_of_sheet = obj_worksheet.max_column
    allure.attach("Maximum Columns that have data in sheet is : "+str(obj_max_column_of_sheet),attachment_type=allure.attachment_type.TEXT)
    list_of_single_row_multiple_columns = []
    # This loop access the data from specific row and all columns data of the specific row and returns a list 
    # of the single row and all columns
    for i in range(1,obj_max_column_of_sheet+1):
        obj_cell = obj_worksheet.cell(row=specific_row_index,column=i)
        str_get_row_values = obj_cell.value
        list_of_single_row_multiple_columns.append(str_get_row_values)
    return list_of_single_row_multiple_columns


"""Author : Sadiya Kotwal
       Description : This method access the excel data from single specific row and single specific columns from that row and column
       Arguments : filepath(filepath='Alfasigma USA, Inc. GP Calc Audit Log 2023-07-31')
                sheetname(Eg: sheetname='C7R Report')
                cell_value(cell_value='A1 or B1' or any cell value whose data is to be acccessed)
        Returns : str_specific_single_value(str_specific_single_value='Returns the single cell value')""" 
def access_single_data_from_specific_cell(filepath,sheetname,cell_value):
    wb = load_workbook(filepath,data_only=True)
    obj_worksheet = wb[sheetname] 
    obj_max_row_of_sheet =   obj_worksheet.max_row
    allure.attach("Maximum Rows that have data in sheet is : "+str(obj_max_row_of_sheet),attachment_type=allure.attachment_type.TEXT)
    obj_max_column_of_sheet = obj_worksheet.max_column
    allure.attach("Maximum Columns that have data in sheet is : "+str(obj_max_column_of_sheet),attachment_type=allure.attachment_type.TEXT)
    str_specific_single_value = obj_worksheet[cell_value].value
    return str_specific_single_value


"""Author : Sadiya Kotwal
       Description : This method access the excel data with key and pair value . Just provide the index for key row and index for column row
       Arguments : filepath(filepath='Alfasigma USA, Inc. GP Calc Audit Log 2023-07-31')
                sheetname(Eg: sheetname='C7R Report')
                fixed_row_index_for_key(fixed_row_index_for_key='8 or 9 or any index of key ))
                fixed_row_index_for_pair(fixed_row_index_for_pair='8 or 9 or any index of  pair value))
        Returns : dict_key_pair(dict_key_pair='Returns the dict value of key and pair for table headers and values')""" 
def access_data_with_key_pair_value_for_headers_with_key_pair(filepath,sheetname,fixed_row_index_for_key,fixed_row_index_for_pair):
    wb = load_workbook(filepath,data_only=True)
    obj_worksheet = wb[sheetname]
    obj_max_row_of_sheet =   obj_worksheet.max_row
    obj_max_column_of_sheet = obj_worksheet.max_column #
    dict_key_pair ={}
    list_keys=[]
    list_pairs = []
    for j in range(1,obj_max_column_of_sheet+1):
        obj_cell = obj_worksheet.cell(row=fixed_row_index_for_key,column=j)
        list_keys.append(obj_cell.value)
    for j in range(1,obj_max_column_of_sheet+1):
        obj_cell = obj_worksheet.cell(row=fixed_row_index_for_pair,column=j)
        list_pairs.append(obj_cell.value)
    for j in range(0,obj_max_column_of_sheet):      
        dict_key_pair[list_keys[j]] = list_pairs[j]
    return dict_key_pair


"""Author : Sadiya Kotwal
       Description : This method access the excel data with key and pair value . This is for the entire sheet
       Arguments : filepath(filepath='Alfasigma USA, Inc. GP Calc Audit Log 2023-07-31')
                sheetname(Eg: sheetname='C7R Report')
                use_header_columns_list(use_header_columns_list='['Name','Calc Id','Price Type','Stage','Approval Submitted By','Approved By','Mark as Delivered By']'))
        Returns : returned_dict(returned_dict='Returns the dict value of key and pair for entire sheet')""" 
def access_data_with_key_pair_value_for_entire_excel_sheet(filepath,sheetname,use_header_columns_list):
    excel_data_df = pandas.read_excel(filepath, sheet_name=sheetname, usecols=use_header_columns_list)
    returned_dict = excel_data_df.to_dict(orient='record')
    allure.attach("Returned Dict value : "+str(returned_dict),attachment_type=allure.attachment_type.TEXT)
    return returned_dict